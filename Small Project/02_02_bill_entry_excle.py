import openpyxl
wb = openpyxl.load_workbook('E:\\Python\\Python My Ch Files\\bill.xlsx')
sh1 = wb['new_sheet_name']

sum = 0
i = 1
product_name_store = []
product_ammount_store = []
print("\n\t********** WELCOME **********")
print("\nEnter q to quit")

while (True):
    product_name = input(f"\nEnter product {i} name : ")
    i += 1

    if (product_name != "q"):

        ammount_input = input(f"Enter {product_name}'s price : ")
        sum = sum + int(ammount_input)
        product_name_store.append(product_name)
        product_ammount_store.append(ammount_input)

        sh1.cell(row=1, column=1, value='product')
        sh1.cell(row=1, column=2, value='price')

        sh1.append([product_name, ammount_input])


    else:
        break

wb.save('E:\\Python\\Python My Ch Files\\bill.xlsx')
print("\n*****************************")
print("Product : Ammount\n")
n = 0
for n in range(i-2):

    print(f"{n+1} {product_name_store[n]} : \t{product_ammount_store[n]}")
    n += 1
print(f"\t Total : {sum}\n")

print("\n*****************************")

print(f"Total Product : {i-2}")
print(f"Your total bill : {sum}")

print("\n© Copyright by Soumojit Shome")

print("\n*****************************")
