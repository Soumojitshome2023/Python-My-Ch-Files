import openpyxl

wb = openpyxl.load_workbook('E:\\Python\\Python My Ch Files\\input_excle.xlsx')
sh1 = wb['new_sheet']
for i in range(2,5):

    pr = input("Enter product : ")
    price = input("Enter price : ")
    sh1.cell(row=1, column=1, value='product')
    sh1.cell(row=1, column=2, value='price')

    sh1.cell(row=i, column=1, value=pr)
    sh1.cell(row=i, column=2, value=price)


wb.save('E:\\Python\\Python My Ch Files\\input_excle.xlsx')
