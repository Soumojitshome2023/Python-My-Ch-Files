import openpyxl

wb = openpyxl.load_workbook('E:\\Python\\Python My Ch Files\\append_excle.xlsx')
sh1 = wb['append_sheet']
for i in range(2,5):

    product_name = input("Enter product name : ")
    price = input("Enter price : ")
    sh1.cell(row=1, column=1, value='product')
    sh1.cell(row=1, column=2, value='price')

    sh1.append([product_name, price])

wb.save('E:\\Python\\Python My Ch Files\\append_excle.xlsx')
